import asyncio
import re
import os
import json
import logging
from datetime import datetime, date

from telegram import Update, InlineKeyboardButton, InlineKeyboardMarkup, ReplyKeyboardMarkup, KeyboardButton
from telegram.ext import (
    Application, CommandHandler, CallbackQueryHandler,
    MessageHandler, filters, ContextTypes
)

# ─── Configuration ───
BOT_TOKEN      = "8194162003:AAFArsa7IIyjGPYselHX7OvGYi83nnXIkwc"
OWNER_ID       = 7095358778
SUPPORT_USER   = "@sadhin8miya"
DATA_FILE      = "tg_checker_data.json"
SESSION_FILE   = "tg_session"
USERS_PER_PAGE = 10

API_ID   = int(os.environ.get("API_ID", "0"))
API_HASH = os.environ.get("API_HASH", "")

logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')
logger = logging.getLogger(__name__)

tg_client = None


# ═══════════════════════════════════════
# ─── Data ───
# ═══════════════════════════════════════

def load_data():
    try:
        with open(DATA_FILE, "r") as f:
            return json.load(f)
    except:
        return {
            "users": {}, "banned": [], "approved": [],
            "settings": {"daily_limit": 0, "open_mode": True, "checker_connected": False},
        }

def save_data(d):
    with open(DATA_FILE, "w") as f:
        json.dump(d, f, indent=2, ensure_ascii=False)

db = load_data()

def is_banned(uid):       return str(uid) in db.get("banned", [])
def is_open_mode():       return db.get("settings", {}).get("open_mode", True)
def is_approved(uid):     return is_open_mode() or str(uid) in db.get("approved", [])
def checker_connected():  return db.get("settings", {}).get("checker_connected", False)

def register_user(update):
    uid = str(update.effective_user.id)
    if uid not in db["users"]:
        db["users"][uid] = {
            "name": update.effective_user.full_name,
            "username": update.effective_user.username or "",
            "joined": datetime.now().isoformat(),
            "checks_total": 0, "checks_today": 0,
            "checks_date": str(date.today()), "custom_limit": -1,
        }
        save_data(db)
    else:
        db["users"][uid]["name"] = update.effective_user.full_name
        db["users"][uid]["username"] = update.effective_user.username or ""

def check_daily_limit(uid):
    u = db["users"].get(uid, {})
    today = str(date.today())
    if u.get("checks_date") != today:
        u["checks_today"] = 0; u["checks_date"] = today; save_data(db)
    cl = u.get("custom_limit", -1)
    if cl == -1:
        gl = db.get("settings", {}).get("daily_limit", 0)
        if gl == 0: return True, -1
        limit = gl
    elif cl == 0: return True, -1
    else: limit = cl
    done = u.get("checks_today", 0)
    return done < limit, max(0, limit - done)

def add_checks(uid, count):
    u = db["users"].get(uid, {})
    today = str(date.today())
    if u.get("checks_date") != today:
        u["checks_today"] = 0; u["checks_date"] = today
    u["checks_today"] = u.get("checks_today", 0) + count
    u["checks_total"] = u.get("checks_total", 0) + count
    save_data(db)


# ═══════════════════════════════════════
# ─── Pyrogram ───
# ═══════════════════════════════════════

async def init_pyrogram():
    global tg_client
    if not API_ID or not API_HASH:
        logger.warning("⚠️ API_ID/API_HASH not set!")
        return
    try:
        from pyrogram import Client
        tg_client = Client(SESSION_FILE, api_id=API_ID, api_hash=API_HASH)
        await tg_client.start()
        db["settings"]["checker_connected"] = True
        save_data(db)
        logger.info("✅ Pyrogram connected!")
    except Exception as e:
        logger.error(f"Pyrogram init error: {e}")
        db["settings"]["checker_connected"] = False
        save_data(db)

async def check_telegram_numbers(numbers):
    global tg_client
    if not tg_client:
        return {n: None for n in numbers}
    from pyrogram.raw import functions, types as raw_types
    results = {}
    for i in range(0, len(numbers), 100):
        batch = numbers[i:i+100]
        contacts = [
            raw_types.InputPhoneContact(
                client_id=idx,
                phone=f"+{re.sub(r'D', '', n)}",
                first_name=f"c{idx}", last_name=""
            ) for idx, n in enumerate(batch)
        ]
        try:
            result = await tg_client.invoke(
                functions.contacts.ImportContacts(contacts=contacts)
            )
            reg = set()
            for u in result.users:
                if hasattr(u, 'phone') and u.phone:
                    reg.add(re.sub(r'\D', '', u.phone))
            for n in batch:
                results[n] = re.sub(r'\D', '', n) in reg
            await tg_client.invoke(
                functions.contacts.DeleteContacts(id=[u.id for u in result.users])
            )
        except Exception as e:
            logger.error(f"Check error: {e}")
            for n in batch: results[n] = None
        await asyncio.sleep(1)
    return results

async def send_tg_code(phone):
    global tg_client
    try:
        from pyrogram import Client
        tg_client = Client(SESSION_FILE, api_id=API_ID, api_hash=API_HASH)
        await tg_client.connect()
        sent = await tg_client.send_code(phone)
        return sent.phone_code_hash
    except Exception as e:
        logger.error(f"Send code error: {e}")
        return ""

async def verify_tg_code(phone, code, phone_hash):
    global tg_client
    try:
        await tg_client.sign_in(phone, phone_hash, code)
        db["settings"]["checker_connected"] = True
        save_data(db)
        return True
    except Exception as e:
        logger.error(f"Sign in error: {e}")
        return False


# ═══════════════════════════════════════
# ─── Keyboards ───
# ═══════════════════════════════════════

def main_kb(uid):
    kb = [
        [KeyboardButton("🔍 Check Numbers"), KeyboardButton("📁 Upload File")],
        [KeyboardButton("📊 My Status"),     KeyboardButton("💬 Support")],
    ]
    if uid == OWNER_ID:
        kb.append([KeyboardButton("⚙️ Admin Panel")])
    return ReplyKeyboardMarkup(kb, resize_keyboard=True)

def admin_menu():
    connected = checker_connected()
    mode      = "🔓 Open" if is_open_mode() else "🔒 Closed"
    return InlineKeyboardMarkup([
        [InlineKeyboardButton(f"📱 Checker: {'🟢 ON' if connected else '🔴 OFF'}",
                              callback_data="adm_checker")],
        [InlineKeyboardButton("👥 Users",        callback_data="adm_users:0"),
         InlineKeyboardButton("🚫 Banned",       callback_data="adm_banned")],
        [InlineKeyboardButton("📢 Broadcast",    callback_data="adm_broadcast"),
         InlineKeyboardButton("✉️ Msg User",     callback_data="adm_msg_user")],
        [InlineKeyboardButton("✅ Approve",       callback_data="adm_approve"),
         InlineKeyboardButton("❌ Unapprove",     callback_data="adm_unapprove"),
         InlineKeyboardButton(f"🔑 {mode}",      callback_data="adm_mode_toggle")],
        [InlineKeyboardButton("⚡ Global Limit",  callback_data="adm_limit"),
         InlineKeyboardButton("👤 User Limit",   callback_data="adm_user_limit")],
        [InlineKeyboardButton("📊 Stats",        callback_data="adm_stats"),
         InlineKeyboardButton("📥 Export Users", callback_data="adm_export")],
        [InlineKeyboardButton("🔙 Close",        callback_data="adm_close")],
    ])


# ═══════════════════════════════════════
# ─── Core Check ───
# ═══════════════════════════════════════

async def do_check(update, context, numbers):
    uid = str(update.effective_user.id)
    if not checker_connected():
        return await update.message.reply_text("❌ Checker connected নেই!\nAdmin এর সাথে যোগাযোগ করুন।")
    can, remaining = check_daily_limit(uid)
    if not can:
        cl = db["users"].get(uid, {}).get("custom_limit", -1)
        gl = db.get("settings", {}).get("daily_limit", 0)
        limit = cl if cl > 0 else gl
        return await update.message.reply_text(
            f"⛔ দৈনিক limit শেষ!\nLimit: *{limit}*/day\nকাল আবার try করুন।",
            parse_mode="Markdown"
        )
    if remaining != -1 and len(numbers) > remaining:
        numbers = numbers[:remaining]
        await update.message.reply_text(f"⚠️ *{len(numbers)}* নম্বর check হবে।", parse_mode="Markdown")

    loading    = await update.message.reply_text(f"⏳ *{len(numbers)}* নম্বর check করছি...", parse_mode="Markdown")
    results    = await check_telegram_numbers(numbers)
    registered = [n for n in numbers if results.get(n) is True]
    fresh      = [n for n in numbers if results.get(n) is False]
    failed     = [n for n in numbers if results.get(n) is None]
    ts         = datetime.now().strftime("%Y%m%d_%H%M%S")
    add_checks(uid, len(numbers))

    await loading.edit_text(
        f"✅ *Check Complete!*\n\n"
        f"📊 Total: *{len(numbers)}*\n"
        f"✅ Telegram আছে: *{len(registered)}*\n"
        f"✨ Fresh (নেই): *{len(fresh)}*\n"
        f"⚠️ Failed: *{len(failed)}*",
        parse_mode="Markdown"
    )
    if registered:
        await update.message.reply_document(
            document="\n".join(registered).encode(),
            filename=f"telegram_registered_{ts}.txt",
            caption=f"✅ *Telegram Registered* — {len(registered)} numbers",
            parse_mode="Markdown"
        )
    if fresh:
        await update.message.reply_document(
            document="\n".join(fresh).encode(),
            filename=f"fresh_{ts}.txt",
            caption=f"✨ *Fresh (No Telegram)* — {len(fresh)} numbers",
            parse_mode="Markdown"
        )
    if failed:
        await update.message.reply_document(
            document="\n".join(failed).encode(),
            filename=f"failed_{ts}.txt",
            caption=f"⚠️ *Failed* — {len(failed)} numbers",
            parse_mode="Markdown"
        )


# ═══════════════════════════════════════
# ─── Handlers ───
# ═══════════════════════════════════════

async def cmd_start(update: Update, context: ContextTypes.DEFAULT_TYPE):
    uid = update.effective_user.id
    if is_banned(uid): return await update.message.reply_text("🚫 You are banned.")
    register_user(update)
    if not is_approved(uid) and uid != OWNER_ID:
        return await update.message.reply_text(
            f"🔒 *Closed mode.*\n\nAccess এর জন্য: {SUPPORT_USER}",
            parse_mode="Markdown"
        )
    status = "🟢 Ready" if checker_connected() else "🔴 Not Ready"
    await update.message.reply_text(
        f"📱 *Telegram Number Checker*\n\n"
        f"🔍 Status: {status}\n\n"
        f"• Text এ নম্বর লিখুন\n"
        f"• 📁 File upload করুন (.txt/.xlsx/.xls)\n"
        f"• Registered/Fresh file পাবেন",
        parse_mode="Markdown",
        reply_markup=main_kb(uid)
    )

async def handle_text(update: Update, context: ContextTypes.DEFAULT_TYPE):
    uid  = update.effective_user.id
    text = update.message.text.strip()
    if is_banned(uid): return
    register_user(update)
    if not is_approved(uid) and uid != OWNER_ID:
        return await update.message.reply_text("🔒 Access denied.")

    sess = context.user_data

    if uid == OWNER_ID:
        state = sess.get("state")

        if state == "broadcast":
            sess.clear()
            users = db.get("users", {})
            sent = fc = 0
            loading = await update.message.reply_text(f"📢 Broadcasting to {len(users)} users...")
            for u_id in users:
                try:
                    await context.bot.send_message(int(u_id), text, parse_mode="Markdown")
                    sent += 1
                except: fc += 1
                await asyncio.sleep(0.05)
            await loading.edit_text(f"✅ Done!\n✅ Sent: {sent}\n❌ Failed: {fc}")
            return

        if state == "msg_user":
            sess.clear()
            parts = text.split("\n", 1)
            if len(parts) < 2:
                return await update.message.reply_text("❌ Format:\n`USER_ID\nMessage`", parse_mode="Markdown")
            try:
                await context.bot.send_message(int(parts[0].strip()), f"📩 *Admin:*\n\n{parts[1]}", parse_mode="Markdown")
                await update.message.reply_text("✅ Sent!")
            except Exception as e:
                await update.message.reply_text(f"❌ Failed: {e}")
            return

        if state == "set_limit":
            sess.clear()
            try:
                limit = int(text.strip())
                db.setdefault("settings", {})["daily_limit"] = limit
                save_data(db)
                await update.message.reply_text(f"✅ Global limit: *{'Unlimited' if limit==0 else f'{limit}/day'}*", parse_mode="Markdown")
            except: await update.message.reply_text("❌ সংখ্যা দিন।")
            return

        if state == "set_user_limit":
            sess.clear()
            parts = text.strip().split()
            if len(parts) != 2:
                return await update.message.reply_text("❌ Format: `USER_ID LIMIT`", parse_mode="Markdown")
            try:
                tid, limit = parts[0], int(parts[1])
                if tid in db.get("users", {}):
                    db["users"][tid]["custom_limit"] = limit
                    save_data(db)
                    lbl = "Unlimited" if limit==0 else f"{limit}/day" if limit>0 else "Global"
                    await update.message.reply_text(f"✅ `{tid}` → *{lbl}*", parse_mode="Markdown")
                else: await update.message.reply_text("❌ User not found.")
            except: await update.message.reply_text("❌ Format: `USER_ID LIMIT`", parse_mode="Markdown")
            return

        if state == "ban_user":
            sess.clear()
            t = text.strip()
            if t not in db.setdefault("banned", []): db["banned"].append(t)
            save_data(db)
            await update.message.reply_text(f"🚫 `{t}` banned.", parse_mode="Markdown")
            return

        if state == "unban_user":
            sess.clear()
            t = text.strip()
            if t in db.get("banned", []): db["banned"].remove(t)
            save_data(db)
            await update.message.reply_text(f"✅ `{t}` unbanned.", parse_mode="Markdown")
            return

        if state == "approve_user":
            sess.clear()
            t = text.strip()
            ap = db.setdefault("approved", [])
            if t not in ap: ap.append(t)
            save_data(db)
            await update.message.reply_text(f"✅ `{t}` approved!", parse_mode="Markdown")
            return

        if state == "unapprove_user":
            sess.clear()
            t = text.strip()
            if t in db.get("approved", []): db["approved"].remove(t)
            save_data(db)
            await update.message.reply_text(f"❌ `{t}` unapproved.", parse_mode="Markdown")
            return

        if state == "checker_phone":
            phone = re.sub(r"\D", "", text.strip())
            sess["state"] = "checker_otp"
            sess["checker_phone"] = phone
            loading = await update.message.reply_text("⏳ OTP পাঠানো হচ্ছে...")
            phone_hash = await send_tg_code(f"+{phone}")
            await loading.delete()
            if not phone_hash:
                sess.clear()
                return await update.message.reply_text("❌ OTP পাঠানো যায়নি। API_ID/API_HASH check করুন।")
            sess["phone_hash"] = phone_hash
            await update.message.reply_text("📱 Telegram OTP code দিন:")
            return

        if state == "checker_otp":
            code  = text.strip().replace(" ", "").replace("-", "")
            phone = sess.get("checker_phone", "")
            ph    = sess.get("phone_hash", "")
            sess.clear()
            loading = await update.message.reply_text("⏳ Connecting...")
            ok = await verify_tg_code(f"+{phone}", code, ph)
            await loading.delete()
            if ok:
                await update.message.reply_text("✅ *Telegram Checker Connected!*", parse_mode="Markdown", reply_markup=main_kb(uid))
            else:
                await update.message.reply_text("❌ OTP ভুল। আবার try করুন।")
            return

    if text == "⚙️ Admin Panel" and uid == OWNER_ID:
        await update.message.reply_text("⚙️ *Admin Panel*", parse_mode="Markdown", reply_markup=admin_menu())
        return

    if text == "🔍 Check Numbers":
        sess["state"] = "waiting_numbers"
        await update.message.reply_text("📝 নম্বর লিখুন বা paste করুন:")
        return

    if text == "📁 Upload File":
        await update.message.reply_text("📁 .txt, .xlsx বা .xls file পাঠান।")
        return

    if text == "📊 My Status":
        u  = db["users"].get(str(uid), {})
        cl = u.get("custom_limit", -1)
        gl = db.get("settings", {}).get("daily_limit", 0)
        _, rem = check_daily_limit(str(uid))
        lbl = "Unlimited" if (cl==0 or (cl==-1 and gl==0)) else f"{rem} remaining"
        uname = f"@{u.get('username')}" if u.get("username") else "(no username)"
        await update.message.reply_text(
            f"📊 *Your Status*\n\n"
            f"👤 {u.get('name','N/A')}\n"
            f"🆔 `{uid}`\n"
            f"📛 {uname}\n"
            f"📊 Total: *{u.get('checks_total',0)}*\n"
            f"📅 Today: *{u.get('checks_today',0)}*\n"
            f"⚡ Limit: *{lbl}*\n"
            f"📅 Joined: {u.get('joined','')[:10]}",
            parse_mode="Markdown"
        )
        return

    if text == "💬 Support":
        await update.message.reply_text(
            f"💬 *Support*\n\n{SUPPORT_USER}",
            parse_mode="Markdown",
            reply_markup=InlineKeyboardMarkup([[
                InlineKeyboardButton("💬 Contact", url=f"https://t.me/{SUPPORT_USER.lstrip('@')}")
            ]])
        )
        return

    if sess.get("state") == "waiting_numbers" or re.search(r'\d{7,15}', text):
        sess.pop("state", None)
        numbers = list(dict.fromkeys(re.findall(r'\d{7,15}', text)))
        if not numbers: return await update.message.reply_text("❌ কোনো valid নম্বর নেই।")
        await do_check(update, context, numbers)

async def handle_file(update: Update, context: ContextTypes.DEFAULT_TYPE):
    uid = update.effective_user.id
    if is_banned(uid): return
    register_user(update)
    if not is_approved(uid) and uid != OWNER_ID:
        return await update.message.reply_text("🔒 Access denied.")
    doc = update.message.document
    if not doc: return
    fname = doc.file_name.lower()
    if not (fname.endswith(".txt") or fname.endswith(".xlsx") or fname.endswith(".xls")):
        return await update.message.reply_text("❌ শুধু .txt, .xlsx বা .xls দিন।")
    loading  = await update.message.reply_text("⏳ File পড়ছি...")
    tmp_path = f"/tmp/{doc.file_id}_{doc.file_name}"
    numbers  = []
    try:
        file = await context.bot.get_file(doc.file_id)
        await file.download_to_drive(tmp_path)
        if fname.endswith(".txt"):
            with open(tmp_path, "r", encoding="utf-8", errors="ignore") as f:
                numbers = re.findall(r'\d{7,15}', f.read())
        elif fname.endswith(".xlsx"):
            from openpyxl import load_workbook
            wb = load_workbook(tmp_path, read_only=True, data_only=True)
            for ws in wb.worksheets:
                for row in ws.iter_rows(values_only=True):
                    for cell in row:
                        if cell:
                            d = re.sub(r"\D", "", str(cell))
                            if 7 <= len(d) <= 15: numbers.append(d)
        elif fname.endswith(".xls"):
            import pandas as pd
            df = pd.read_excel(tmp_path, engine="xlrd", dtype=str, header=None)
            for col in df.columns:
                for val in df[col].dropna():
                    d = re.sub(r"\D", "", str(val))
                    if 7 <= len(d) <= 15: numbers.append(d)
        os.remove(tmp_path)
    except Exception as e:
        await loading.delete()
        return await update.message.reply_text(f"❌ File error: {e}")
    numbers = list(dict.fromkeys(numbers))
    await loading.delete()
    if not numbers: return await update.message.reply_text("❌ কোনো নম্বর পাওয়া যায়নি।")
    await do_check(update, context, numbers)


# ═══════════════════════════════════════
# ─── Admin Callbacks ───
# ═══════════════════════════════════════

async def cb_handler(update: Update, context: ContextTypes.DEFAULT_TYPE):
    query = update.callback_query
    uid   = update.effective_user.id
    data  = query.data
    try: await query.answer()
    except: pass

    if data == "cancel":
        context.user_data.clear()
        try: await query.edit_message_text("❌ Cancelled")
        except: pass
        return

    if data == "adm_close":
        try: await query.delete_message()
        except: pass
        return

    if uid != OWNER_ID: return

    if data == "adm_back":
        await query.edit_message_text("⚙️ *Admin Panel*", parse_mode="Markdown", reply_markup=admin_menu())

    elif data == "adm_checker":
        connected = checker_connected()
        await query.edit_message_text(
            f"📱 *Telegram Checker*\n\nStatus: *{'🟢 Connected' if connected else '🔴 Disconnected'}*",
            parse_mode="Markdown",
            reply_markup=InlineKeyboardMarkup([
                [InlineKeyboardButton("📱 Connect", callback_data="adm_checker_connect")],
                [InlineKeyboardButton("🔌 Disconnect", callback_data="adm_checker_disconnect")],
                [InlineKeyboardButton("🔙 Back", callback_data="adm_back")],
            ])
        )

    elif data == "adm_checker_connect":
        context.user_data["state"] = "checker_phone"
        await query.edit_message_text(
            "📱 Telegram নম্বর দিন:\nExample: `8801712345678`",
            parse_mode="Markdown",
            reply_markup=InlineKeyboardMarkup([[InlineKeyboardButton("❌ Cancel", callback_data="cancel")]])
        )

    elif data == "adm_checker_disconnect":
        global tg_client
        try:
            if tg_client: await tg_client.stop()
        except: pass
        tg_client = None
        db["settings"]["checker_connected"] = False
        save_data(db)
        await query.answer("🔌 Disconnected", show_alert=True)
        await query.edit_message_text("⚙️ *Admin Panel*", parse_mode="Markdown", reply_markup=admin_menu())

    elif data == "adm_stats":
        users = db.get("users", {})
        total = sum(u.get("checks_total", 0) for u in users.values())
        gl    = db.get("settings", {}).get("daily_limit", 0)
        await query.edit_message_text(
            f"📊 *Statistics*\n\n"
            f"👥 Users: *{len(users)}*\n"
            f"🚫 Banned: *{len(db.get('banned',[]))}*\n"
            f"✅ Approved: *{len(db.get('approved',[]))}*\n"
            f"🔑 Mode: *{'🔓 Open' if is_open_mode() else '🔒 Closed'}*\n"
            f"🔢 Total Checks: *{total}*\n"
            f"⚡ Daily Limit: *{'Unlimited' if gl==0 else gl}*\n"
            f"📱 Checker: *{'🟢 Connected' if checker_connected() else '🔴 Disconnected'}*",
            parse_mode="Markdown",
            reply_markup=InlineKeyboardMarkup([[InlineKeyboardButton("🔙 Back", callback_data="adm_back")]])
        )

    elif data == "adm_export":
        users    = db.get("users", {})
        banned   = db.get("banned", [])
        approved = db.get("approved", [])
        lines    = [
            "=" * 60,
            "TELEGRAM CHECKER BOT - USER EXPORT",
            f"Date: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}",
            f"Total Users: {len(users)}",
            "=" * 60, ""
        ]
        for uid_str, u in sorted(users.items(), key=lambda x: x[1].get("checks_total", 0), reverse=True):
            uname = u.get("username", "")
            cl    = u.get("custom_limit", -1)
            lbl   = "Global" if cl==-1 else ("Unlimited" if cl==0 else f"{cl}/day")
            lines += [
                f"👤 Name     : {u.get('name','N/A')}",
                f"🆔 User ID  : {uid_str}",
                f"📛 Username : @{uname}" if uname else "📛 Username : (not set)",
                f"📅 Joined   : {u.get('joined','')[:10]}",
                f"📊 Total    : {u.get('checks_total',0)} checks",
                f"📅 Today    : {u.get('checks_today',0)} checks",
                f"⚡ Limit    : {lbl}",
            ]
            if uid_str in banned:   lines.append("🚫 Status   : BANNED")
            if uid_str in approved: lines.append("✅ Status   : APPROVED")
            lines.append("-" * 40)
        ts = datetime.now().strftime("%Y%m%d_%H%M%S")
        await context.bot.send_document(
            chat_id=uid,
            document="\n".join(lines).encode("utf-8"),
            filename=f"users_export_{ts}.txt",
            caption=f"📥 *User Export*\n👥 Total: *{len(users)}* users",
            parse_mode="Markdown"
        )
        await query.answer("✅ File sent!", show_alert=True)

    elif data.startswith("adm_users:"):
        page  = int(data.split(":")[1])
        users = list(db.get("users", {}).items())
        total = len(users)
        chunk = users[page*USERS_PER_PAGE:(page+1)*USERS_PER_PAGE]
        text  = f"👥 *Users* ({total}) — Page {page+1}\n\n"
        for u_id, u in chunk:
            bm    = "🚫" if u_id in db.get("banned",[]) else ""
            am    = "✅" if u_id in db.get("approved",[]) else ""
            uname = f"@{u['username']}" if u.get("username") else "(no username)"
            text += f"{bm}{am} {u.get('name','?')} | {uname}\n🆔 `{u_id}` | 📊 {u.get('checks_total',0)}\n\n"
        nav = []
        if page > 0: nav.append(InlineKeyboardButton("◀️", callback_data=f"adm_users:{page-1}"))
        if (page+1)*USERS_PER_PAGE < total: nav.append(InlineKeyboardButton("▶️", callback_data=f"adm_users:{page+1}"))
        buttons = []
        if nav: buttons.append(nav)
        buttons += [
            [InlineKeyboardButton("🚫 Ban",   callback_data="adm_do_ban"),
             InlineKeyboardButton("✅ Unban", callback_data="adm_do_unban")],
            [InlineKeyboardButton("🔙 Back",  callback_data="adm_back")],
        ]
        await query.edit_message_text(text, parse_mode="Markdown", reply_markup=InlineKeyboardMarkup(buttons))

    elif data == "adm_banned":
        banned = db.get("banned", [])
        text   = f"🚫 *Banned* ({len(banned)})\n\n"
        for b in banned[:20]:
            text += f"• `{b}` — {db['users'].get(b,{}).get('name','?')}\n"
        await query.edit_message_text(text or "কেউ নেই।", parse_mode="Markdown",
            reply_markup=InlineKeyboardMarkup([
                [InlineKeyboardButton("✅ Unban", callback_data="adm_do_unban")],
                [InlineKeyboardButton("🔙 Back",  callback_data="adm_back")]
            ]))

    elif data == "adm_do_ban":
        context.user_data["state"] = "ban_user"
        await query.edit_message_text("🚫 Ban করতে User ID দিন:",
            reply_markup=InlineKeyboardMarkup([[InlineKeyboardButton("❌ Cancel", callback_data="cancel")]]))

    elif data == "adm_do_unban":
        context.user_data["state"] = "unban_user"
        await query.edit_message_text("✅ Unban করতে User ID দিন:",
            reply_markup=InlineKeyboardMarkup([[InlineKeyboardButton("❌ Cancel", callback_data="cancel")]]))

    elif data == "adm_broadcast":
        context.user_data["state"] = "broadcast"
        await query.edit_message_text("📢 Broadcast message লিখুন: (Markdown supported)",
            reply_markup=InlineKeyboardMarkup([[InlineKeyboardButton("❌ Cancel", callback_data="cancel")]]))

    elif data == "adm_msg_user":
        context.user_data["state"] = "msg_user"
        await query.edit_message_text("✉️ Format:\n`USER_ID\nMessage`", parse_mode="Markdown",
            reply_markup=InlineKeyboardMarkup([[InlineKeyboardButton("❌ Cancel", callback_data="cancel")]]))

    elif data == "adm_approve":
        context.user_data["state"] = "approve_user"
        mode = "🔓 Open" if is_open_mode() else "🔒 Closed"
        await query.edit_message_text(
            f"✅ *Approve User*\nMode: *{mode}*\n\nUser ID দিন:",
            parse_mode="Markdown",
            reply_markup=InlineKeyboardMarkup([[InlineKeyboardButton("❌ Cancel", callback_data="cancel")]]))

    elif data == "adm_unapprove":
        context.user_data["state"] = "unapprove_user"
        approved = db.get("approved", [])
        lst = "\n".join([f"• `{a}`" for a in approved[:15]]) or "কেউ নেই"
        await query.edit_message_text(f"❌ Approved:\n{lst}\n\nID দিন:",
            parse_mode="Markdown",
            reply_markup=InlineKeyboardMarkup([[InlineKeyboardButton("❌ Cancel", callback_data="cancel")]]))

    elif data == "adm_mode_toggle":
        db.setdefault("settings", {})["open_mode"] = not is_open_mode()
        save_data(db)
        mode = "🔓 Open" if is_open_mode() else "🔒 Closed"
        await query.answer(f"Mode: {mode}", show_alert=True)
        await query.edit_message_text("⚙️ *Admin Panel*", parse_mode="Markdown", reply_markup=admin_menu())

    elif data == "adm_limit":
        current = db.get("settings", {}).get("daily_limit", 0)
        context.user_data["state"] = "set_limit"
        await query.edit_message_text(
            f"⚡ *Global Limit*\nCurrent: *{'Unlimited' if current==0 else current}*\n\n0 = unlimited:",
            parse_mode="Markdown",
            reply_markup=InlineKeyboardMarkup([[InlineKeyboardButton("❌ Cancel", callback_data="cancel")]]))

    elif data == "adm_user_limit":
        context.user_data["state"] = "set_user_limit"
        await query.edit_message_text(
            "👤 *User Limit*\n\nFormat: `USER_ID LIMIT`\n\n"
            "`USER_ID 500` → 500/day\n`USER_ID 0` → Unlimited\n`USER_ID -1` → Global",
            parse_mode="Markdown",
            reply_markup=InlineKeyboardMarkup([[InlineKeyboardButton("❌ Cancel", callback_data="cancel")]]))


# ═══════════════════════════════════════
# ─── Main ───
# ═══════════════════════════════════════

def main():
    logger.info("📱 Telegram Checker Bot Starting...")
    app = Application.builder().token(BOT_TOKEN).build()
    app.add_handler(CommandHandler("start", cmd_start))
    app.add_handler(CallbackQueryHandler(cb_handler))
    app.add_handler(MessageHandler(filters.Document.ALL, handle_file))
    app.add_handler(MessageHandler(filters.TEXT & ~filters.COMMAND, handle_text))

    async def post_init(application):
        await init_pyrogram()

    app.post_init = post_init
    app.run_polling(allowed_updates=["message", "callback_query"])

if __name__ == "__main__":
    main()
